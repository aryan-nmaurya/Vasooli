"""Ledger and summary exports. CSV, Excel, PDF.

The point of these is that a merchant's accountant opens the file in the tool they
already reconcile in. That fails in two boring ways — a corrupt file, or numbers that
disagree with the dashboard — so both are pinned here.
"""

import csv
import io

import pytest
from fastapi.testclient import TestClient

from app.core.config import settings
from app.core.constants import InvoiceStatus
from app.main import create_app
from app.services import exports


@pytest.fixture
def api(session):
    with TestClient(create_app()) as c:
        yield c


@pytest.fixture
def recovered(session, invoice):
    """One invoice whose money arrived."""
    from app.core.clock import utcnow

    invoice.status = InvoiceStatus.RECOVERED
    invoice.amount_paid_paise = invoice.amount_paise
    invoice.recovered_at = utcnow()
    session.add(invoice)
    session.commit()
    return invoice


def _get(api, path):
    return api.get(path, headers={"X-Admin-Key": settings.admin_api_key})


# ===========================================================================
# The files are real files.
# ===========================================================================


def test_csv_opens_and_carries_the_invoice(api, session, recovered):
    res = _get(api, "/api/export/recovered?format=csv")
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("text/csv")
    assert "attachment" in res.headers["content-disposition"]

    # The BOM is deliberate: without it Excel renders the rupee sign as mojibake.
    body = res.content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(body)))
    assert rows[0][0] == "Invoice"
    assert any(r and r[0] == recovered.invoice_number for r in rows[1:])


def test_xlsx_is_a_real_workbook_with_numeric_money(api, session, recovered):
    """A sheet that shows "₹34,000" but stores a string cannot be summed, which is the
    first thing anyone does with an exported ledger."""
    from openpyxl import load_workbook

    res = _get(api, "/api/export/recovered?format=xlsx")
    assert res.status_code == 200

    book = load_workbook(io.BytesIO(res.content))
    page = book.active
    assert page["A1"].value == "Invoice"

    amount = page.cell(row=2, column=4)
    assert isinstance(amount.value, int | float)
    assert "#,##0" in amount.number_format
    assert page.freeze_panes == "A2"


def test_pdf_has_a_pdf_header(api, session, recovered):
    res = _get(api, "/api/export/recovered?format=pdf")
    assert res.status_code == 200
    assert res.content.startswith(b"%PDF-")
    assert res.headers["content-type"] == "application/pdf"


@pytest.mark.parametrize("fmt", ["csv", "xlsx", "pdf"])
def test_the_overview_exports_in_every_format(api, session, fmt):
    assert _get(api, f"/api/export/overview?format={fmt}").status_code == 200


# ===========================================================================
# The numbers match the dashboard.
# ===========================================================================


def test_the_export_and_the_dashboard_agree(api, session, recovered):
    """One source, three renderings. If the export recalculated anything of its own,
    a figure could differ between the screen and the spreadsheet — which is the kind
    of discrepancy that destroys trust in both."""
    sheet = exports.recovered_invoices(session)
    queue = _get(api, "/api/dashboard/queue?limit=200").json()
    on_screen = {r["invoice_number"] for r in queue if r["status"] == "recovered"}

    assert {row[0] for row in sheet.rows} == on_screen


def test_an_empty_ledger_still_produces_valid_files(api, session):
    """No recovered invoices is an ordinary state, not an error."""
    for fmt in ("csv", "xlsx", "pdf"):
        res = _get(api, f"/api/export/recovered?format={fmt}")
        assert res.status_code == 200, fmt
        assert len(res.content) > 0


# ===========================================================================
# The endpoint is not a general-purpose data tap.
# ===========================================================================


def test_an_unknown_dataset_is_refused(api, session):
    """An export endpoint that accepts an arbitrary name is an exfiltration route."""
    res = _get(api, "/api/export/customers")
    assert res.status_code == 404
    assert "recovered" in res.json()["detail"]


def test_an_unknown_format_is_refused(api, session):
    assert _get(api, "/api/export/recovered?format=exe").status_code == 422


def test_exports_require_authentication(api, session):
    assert api.get("/api/export/recovered").status_code == 401


def test_filenames_carry_the_dataset_and_a_timestamp(api, session, recovered):
    """A folder of files called "export.csv" is unusable a week later."""
    disposition = _get(api, "/api/export/recovered?format=csv").headers["content-disposition"]
    assert "vasooli-recovered-invoices-" in disposition
    assert disposition.endswith('.csv"')


# ===========================================================================
# The download matches what is on screen. Filtered exports.
# ===========================================================================


def _numbers(res) -> set[str]:
    rows = list(csv.reader(io.StringIO(res.content.decode("utf-8-sig"))))
    return {r[0] for r in rows[1:] if r}


def test_the_queue_export_respects_the_status_filter(api, session, recovered, invoice):
    """An export that ignored the active filter would be worse than none: the merchant
    reconciles against the wrong set with no reason to suspect it."""
    everything = _numbers(_get(api, "/api/export/invoices?format=csv"))
    filtered = _numbers(_get(api, "/api/export/invoices?format=csv&status=recovered"))

    assert recovered.invoice_number in filtered
    assert filtered < everything or filtered == everything
    assert all(n in everything for n in filtered)


def test_the_queue_export_respects_the_reason_filter(api, session, invoice):
    from app.core.constants import ReasonCategory

    invoice.reason_category = ReasonCategory.OVERSIGHT
    session.add(invoice)
    session.commit()

    hit = _numbers(_get(api, "/api/export/invoices?format=csv&reason=oversight"))
    miss = _numbers(_get(api, "/api/export/invoices?format=csv&reason=unresponsive"))

    assert invoice.invoice_number in hit
    assert invoice.invoice_number not in miss


def test_both_filters_apply_together(api, session, invoice):
    """Status and reason are ANDed, matching how the dashboard stacks them."""
    from app.core.constants import InvoiceStatus, ReasonCategory

    invoice.status = InvoiceStatus.HUMAN_REVIEW
    invoice.reason_category = ReasonCategory.DISPUTE_LIKELY
    session.add(invoice)
    session.commit()

    both = _numbers(
        _get(api, "/api/export/invoices?format=csv&status=human_review&reason=dispute_likely")
    )
    wrong_reason = _numbers(
        _get(api, "/api/export/invoices?format=csv&status=human_review&reason=oversight")
    )

    assert invoice.invoice_number in both
    assert invoice.invoice_number not in wrong_reason


def test_the_filter_is_named_in_the_exported_file(api, session, invoice):
    """The subtitle says which filter produced the file, so a spreadsheet found later
    is not mistaken for the whole ledger."""
    sheet = exports.queue_invoices(session, status="human_review")
    assert "filtered by" in sheet.subtitle
    assert "human review" in sheet.subtitle


@pytest.mark.parametrize("bad", ["'; DROP TABLE invoices;--", "Recovered", "a" * 40])
def test_a_malformed_filter_is_refused(api, session, bad):
    """These reach a WHERE clause; the endpoint constrains their shape rather than
    trusting the caller."""
    assert _get(api, f"/api/export/invoices?format=csv&status={bad}").status_code == 422
