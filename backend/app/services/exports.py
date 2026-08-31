"""Ledger and summary exports. CSV, Excel, PDF.

A merchant's accountant does not log into dashboards; they want the numbers in the
tool they already reconcile in. These render the same data three ways from one source,
so a figure can never differ between formats — the rows are built once and each
renderer only decides how to draw them.

Money is formatted for humans in the display column and kept as exact integer paise in
the numeric one. A spreadsheet that shows "₹34,000" but sums a string is worse than one
that shows nothing, so the amount columns are written as real numbers Excel can total.
"""

import csv
import io
import uuid
from dataclasses import dataclass

from sqlmodel import Session, select

from app.core.clock import now_ist
from app.core.money import format_inr
from app.models import Customer, Invoice, PaymentLink
from app.services.demo_scope import demo_invoices

#: One row of an export, plus the metadata a renderer needs to lay it out.
Row = list[object]


@dataclass(frozen=True)
class Sheet:
    """A rendered table, independent of output format."""

    title: str
    subtitle: str
    headers: list[str]
    rows: list[Row]
    #: Column indexes holding money, as rupees. Excel formats these as currency and
    #: the PDF right-aligns them.
    money_columns: tuple[int, ...] = ()

    @property
    def filename_stem(self) -> str:
        stamp = now_ist().strftime("%Y%m%d-%H%M")
        slug = self.title.lower().replace(" ", "-")
        return f"vasooli-{slug}-{stamp}"


def _paise_to_rupees(paise: int) -> float:
    """Rupees for a spreadsheet cell.

    The only place this codebase converts money to a float, and it is deliberate: a
    spreadsheet cell is a float either way, so the choice is between converting here
    with a known rounding rule or letting Excel parse a string and guess. Never fed
    back into any calculation — the authoritative value stays integer paise in the
    database.
    """
    return round(paise / 100, 2)


def _lookups(
    session: Session, invoices: list[Invoice]
) -> tuple[dict[uuid.UUID, Customer], dict[uuid.UUID, PaymentLink]]:
    """Fetch only the customers and payment links the exported rows actually reference.

    Both builders used to run a bare `select(Customer)` and `select(PaymentLink)` and
    hold every row in a dict, to resolve a name and a URL per invoice. Nothing leaked —
    the dicts are only ever read by id, and row-level security scopes them to the
    tenant — but the cost is the whole table in memory on every export, for a filtered
    view that may be twenty rows. It grows with the merchant's history rather than with
    the size of the download, which is the wrong axis.

    Chunked because a query is not allowed to grow without bound either: Postgres has a
    parameter ceiling, and one enormous IN list is its own failure.
    """
    if not invoices:
        return {}, {}

    customer_ids = {inv.customer_id for inv in invoices}
    invoice_ids = [inv.id for inv in invoices]

    names: dict[uuid.UUID, Customer] = {}
    for chunk in _chunked(sorted(customer_ids, key=str), _ID_CHUNK):
        for row in session.exec(select(Customer).where(Customer.id.in_(chunk))).all():  # type: ignore[attr-defined]
            names[row.id] = row

    links: dict[uuid.UUID, PaymentLink] = {}
    for chunk in _chunked(invoice_ids, _ID_CHUNK):
        for row in session.exec(
            select(PaymentLink).where(PaymentLink.invoice_id.in_(chunk))  # type: ignore[attr-defined]
        ).all():
            links[row.invoice_id] = row

    return names, links


#: Comfortably under Postgres' bind-parameter ceiling, with room for the rest of the
#: statement.
_ID_CHUNK = 1000


def _chunked(items: list, size: int):
    for start in range(0, len(items), size):
        yield items[start : start + size]


def recovered_invoices(session: Session, **_: object) -> Sheet:
    """Every invoice whose money actually arrived."""
    from app.core.constants import InvoiceStatus

    invoices = list(
        session.exec(
            demo_invoices()
            .where(Invoice.status == InvoiceStatus.RECOVERED)
            .order_by(Invoice.recovered_at.desc())  # type: ignore[attr-defined]
        ).all()
    )
    names, links = _lookups(session, invoices)

    rows: list[Row] = []
    for inv in invoices:
        customer = names.get(inv.customer_id)
        link = links.get(inv.id)
        days_to_recovery = (
            (inv.recovered_at.date() - inv.due_at.date()).days
            if inv.recovered_at and inv.due_at
            else None
        )
        rows.append(
            [
                inv.invoice_number,
                customer.name if customer else "—",
                customer.email if customer else "—",
                _paise_to_rupees(inv.amount_paise),
                _paise_to_rupees(inv.amount_paid_paise),
                inv.issued_at.date().isoformat() if inv.issued_at else "",
                inv.due_at.date().isoformat() if inv.due_at else "",
                inv.recovered_at.date().isoformat() if inv.recovered_at else "",
                days_to_recovery if days_to_recovery is not None else "",
                inv.reminders_sent,
                str(inv.reason_category) if inv.reason_category else "",
                link.razorpay_payment_link_id if link else "",
            ]
        )

    total = sum(i.amount_paid_paise for i in invoices)
    return Sheet(
        title="Recovered invoices",
        subtitle=f"{len(invoices)} invoices · {format_inr(total)} recovered",
        headers=[
            "Invoice",
            "Customer",
            "Email",
            "Amount (₹)",
            "Paid (₹)",
            "Issued",
            "Due",
            "Recovered",
            "Days to recovery",
            "Reminders sent",
            "Reason",
            "Razorpay link",
        ],
        rows=rows,
        money_columns=(3, 4),
    )


def overview_summary(session: Session, *, days: int = 30, **_: object) -> Sheet:
    """The dashboard's headline numbers, as a two-column sheet."""
    from datetime import timedelta

    from app.core.clock import utcnow
    from app.services.metrics import compute_metrics

    # `since`, matching what the dashboard's own overview endpoint passes, so the
    # exported figures are the same ones on screen rather than a parallel calculation.
    m = compute_metrics(session, since=utcnow() - timedelta(days=days)).as_dict()
    pairs = [
        ("Total overdue", m["total_overdue_display"]),
        ("Recovered", m["recovered_display"]),
        ("Recovery rate (by value)", m["recovery_rate_display"]),
        ("Average days to recovery", str(m["avg_days_to_recovery"] or "—")),
        ("Automation rate", m["automation_rate_display"]),
        ("Invoices total", str(m["invoices_total"])),
        ("Invoices recovered", str(m["invoices_recovered"])),
        ("Needing a human", str(m["invoices_in_human_review"])),
        ("Active promises", str(m["active_promises"])),
        ("Broken promises", str(m["broken_promises"])),
    ]
    pairs += [(f"Status · {k}", str(v)) for k, v in sorted(m["counts_by_status"].items())]
    pairs += [(f"Reason · {k}", str(v)) for k, v in sorted(m["counts_by_reason"].items())]

    return Sheet(
        title="Recovery overview",
        subtitle=f"Last {days} days · generated {now_ist().strftime('%d %b %Y, %H:%M')} IST",
        headers=["Metric", "Value"],
        rows=[[label, value] for label, value in pairs],
    )


# --------------------------------------------------------------------------
# Renderers.
# --------------------------------------------------------------------------


def to_csv(sheet: Sheet) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(sheet.headers)
    writer.writerows(sheet.rows)
    # BOM so Excel opens a UTF-8 CSV with the rupee sign intact rather than as mojibake.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(sheet: Sheet) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    page = book.active
    page.title = sheet.title[:31]  # Excel's hard limit on sheet names

    page.append(sheet.headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in page[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in sheet.rows:
        page.append(row)

    # Real currency cells, so a merchant can sum the column instead of retyping it.
    for index in sheet.money_columns:
        letter = get_column_letter(index + 1)
        for cell in page[letter][1:]:
            cell.number_format = "#,##0.00"

    for index, header in enumerate(sheet.headers, start=1):
        longest = max([len(str(header))] + [len(str(r[index - 1])) for r in sheet.rows] or [0])
        page.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 42)

    page.freeze_panes = "A2"

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def to_pdf(sheet: Sheet) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    # Landscape: the recovered-invoice sheet is twelve columns and portrait forces
    # either a font nobody can read or a table that runs off the page.
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4) if len(sheet.headers) > 4 else A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=sheet.title,
        author="Vasooli",
    )

    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"<b>{sheet.title}</b>", styles["Title"]),
        Paragraph(sheet.subtitle, styles["Normal"]),
        Spacer(1, 6 * mm),
    ]

    data = [sheet.headers] + [[str(c) for c in row] for row in sheet.rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            # Money reads wrong left-aligned; the eye compares magnitudes down a column.
            + [("ALIGN", (i, 1), (i, -1), "RIGHT") for i in sheet.money_columns]
        )
    )
    story.append(table)

    generated = now_ist().strftime("%d %b %Y, %H:%M")
    story += [
        Spacer(1, 5 * mm),
        Paragraph(
            f"<font size=7 color='#6B7280'>Generated by Vasooli · {generated} IST · "
            f"Razorpay test mode</font>",
            styles["Normal"],
        ),
    ]
    doc.build(story)
    return out.getvalue()


RENDERERS = {
    "csv": (to_csv, "text/csv; charset=utf-8", "csv"),
    "xlsx": (
        to_xlsx,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "pdf": (to_pdf, "application/pdf", "pdf"),
}


def render(sheet: Sheet, fmt: str) -> tuple[bytes, str, str]:
    """Bytes, content type, and filename for one sheet in one format."""
    if fmt not in RENDERERS:
        raise ValueError(f"unsupported format {fmt!r}")
    renderer, content_type, extension = RENDERERS[fmt]
    return renderer(sheet), content_type, f"{sheet.filename_stem}.{extension}"


def queue_invoices(
    session: Session,
    *,
    status: str | None = None,
    reason: str | None = None,
    **_: object,
) -> Sheet:
    """The recovery queue, filtered exactly as the dashboard filters it.

    Takes the same two filters the overview screen exposes, so what downloads is what
    the merchant is looking at. An export that silently ignored the active filter
    would be worse than no export — they would reconcile against the wrong set and
    have no reason to suspect it.
    """
    from app.core.constants import InvoiceStatus  # noqa: F401  (documents the vocabulary)

    query = demo_invoices()
    if status:
        query = query.where(Invoice.status == status)
    if reason:
        query = query.where(Invoice.reason_category == reason)

    invoices = list(session.exec(query).all())
    # Same ordering as the dashboard: largest outstanding first, because that is where
    # a merchant's attention is worth the most.
    invoices.sort(key=lambda i: i.outstanding_paise, reverse=True)

    names, links = _lookups(session, invoices)

    rows: list[Row] = []
    for inv in invoices:
        customer = names.get(inv.customer_id)
        link = links.get(inv.id)
        rows.append(
            [
                inv.invoice_number,
                customer.name if customer else "—",
                customer.email if customer else "—",
                _paise_to_rupees(inv.amount_paise),
                _paise_to_rupees(inv.amount_paid_paise),
                _paise_to_rupees(inv.outstanding_paise),
                inv.days_overdue,
                str(inv.status),
                str(inv.reason_category) if inv.reason_category else "",
                inv.reminders_sent,
                inv.issued_at.date().isoformat() if inv.issued_at else "",
                inv.due_at.date().isoformat() if inv.due_at else "",
                link.short_url if link else "",
            ]
        )

    applied = [f for f in (status, reason) if f]
    outstanding = sum(i.outstanding_paise for i in invoices)
    subtitle = f"{len(invoices)} invoices · {format_inr(outstanding)} outstanding"
    if applied:
        subtitle += " · filtered by " + ", ".join(a.replace("_", " ") for a in applied)

    return Sheet(
        title="Recovery queue",
        subtitle=subtitle,
        headers=[
            "Invoice",
            "Customer",
            "Email",
            "Amount (₹)",
            "Paid (₹)",
            "Outstanding (₹)",
            "Days overdue",
            "Status",
            "Reason",
            "Reminders sent",
            "Issued",
            "Due",
            "Payment link",
        ],
        rows=rows,
        money_columns=(3, 4, 5),
    )
